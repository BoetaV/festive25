from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse_lazy
from django.views.generic import TemplateView, ListView, CreateView, UpdateView, DeleteView
from django.views import View
from django.http import HttpResponse, JsonResponse
from django.db import transaction
from django.db.models import Count, Q, Sum, Case, When, Value, CharField
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.contrib.auth.decorators import login_required # Ensure this is here
from django.template.loader import render_to_string
from django.contrib.staticfiles import finders
from django.contrib import messages # Ensure this is here
from datetime import date, timedelta
import json

# --- Third-Party Library Imports ---
import openpyxl # Ensure this is here
from openpyxl.styles import Font, Alignment # Ensure this is here
from weasyprint import HTML, CSS

# --- Local App Imports ---
from .models import Delivery, Baby
from .forms import DeliveryForm, BabyFormSet, DashboardReportFilterForm, NilReportFilterForm # Ensure NilReportFilterForm is defined
from .data import LOCATION_DATA, DISTRICT_CHOICES, FACILITY_TYPES
from accounts.models import Profile # <--- Correct Import for your Profile model
from django.contrib.auth import get_user_model # To get the active User model

User = get_user_model() # Define User here for consistency

# ==========================================================
# HELPER FUNCTIONS & PERMISSION MIXINS
# ==========================================================
def can_modify_data(user):
    """Permission check for users who can create/edit/delete data."""
    return user.is_authenticated and not user.groups.filter(name='ProvinceUser').exists()

class DataEditorRequiredMixin(UserPassesTestMixin):
    """Mixin to apply the can_modify_data permission check to views."""
    def test_func(self):
        return can_modify_data(self.request.user)

# ==========================================================
# DASHBOARD VIEW
# ==========================================================
from django.views.generic import TemplateView
from django.db.models import Count, Q, Sum
from datetime import date, timedelta
import json

class LandingPageView(TemplateView):
    template_name = 'landing.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # --------------------------------------------------
        # 1. BASE QUERYSETS (ACTIVE DATA ONLY)
        # --------------------------------------------------
            deliveries_qs = Delivery.objects.all()
            babies_base_qs = Baby.objects.all()

        # --------------------------------------------------
        # 2. GET FILTER VALUES FROM REQUEST
        # --------------------------------------------------
        selected_date = self.request.GET.get('report_date')
        selected_district = self.request.GET.get('district')
        selected_municipality = self.request.GET.get('local_municipality')
        selected_facility = self.request.GET.get('facility')

        if selected_date:
            deliveries_qs = deliveries_qs.filter(report_date=selected_date)
        if selected_district:
            deliveries_qs = deliveries_qs.filter(district=selected_district)
        if selected_municipality:
            deliveries_qs = deliveries_qs.filter(local_municipality=selected_municipality)
        if selected_facility:
            deliveries_qs = deliveries_qs.filter(facility=selected_facility)

        # --------------------------------------------------
        # 3. BABIES QUERYSET (SAFE & DISTINCT)
        # --------------------------------------------------
        babies_qs = babies_base_qs.filter(
            delivery__in=deliveries_qs,
            delivery__no_births_to_report=False
        ).distinct()

        # --------------------------------------------------
        # 4. KPI CARDS
        # --------------------------------------------------
        total_births = babies_qs.count()
        total_males = babies_qs.filter(gender='Male').count()
        total_females = babies_qs.filter(gender='Female').count()
        total_nil_reports = deliveries_qs.filter(no_births_to_report=True).count()

        # --------------------------------------------------
        # 5. SUMMARY TABLE CONFIGURATION
        # --------------------------------------------------
        if selected_facility:
            group_by = 'facility'
            title = f'Births in {selected_facility}'
            header = 'Facility'
            footer = selected_facility
        elif selected_municipality:
            group_by = 'facility'
            title = f'Births per Facility in {selected_municipality}'
            header = 'Facility'
            footer = selected_municipality
        elif selected_district:
            group_by = 'local_municipality'
            title = f'Births per Local Municipality in {selected_district}'
            header = 'Local Municipality'
            footer = selected_district
        else:
            group_by = 'district'
            title = 'Births per District'
            header = 'District'
            footer = 'Eastern Cape'

        summary_data = deliveries_qs.filter(
            no_births_to_report=False
        ).values(group_by).annotate(
            total_babies=Count('babies', distinct=True),
            male_count=Count('babies', filter=Q(babies__gender='Male'), distinct=True),
            female_count=Count('babies', filter=Q(babies__gender='Female'), distinct=True),
        ).order_by(group_by)

        # --------------------------------------------------
        # 6. AGE GROUP SUMMARY
        # --------------------------------------------------
        age_group_labels = ["10-14 yrs", "15-19 yrs", "20-35 yrs", "35+ yrs"]
        age_group_summary = {label: {'male_count': 0, 'female_count': 0, 'total': 0} for label in age_group_labels}
        today = date.today()

        deliveries_for_age = deliveries_qs.filter(
            mother_dob__isnull=False,
            no_births_to_report=False
        ).prefetch_related('babies')

        for delivery in deliveries_for_age:
            age = today.year - delivery.mother_dob.year - (
                (today.month, today.day) < (delivery.mother_dob.month, delivery.mother_dob.day)
            )

            if 10 <= age <= 14:
                group = "10-14 yrs"
            elif 15 <= age <= 19:
                group = "15-19 yrs"
            elif 20 <= age <= 35:
                group = "20-35 yrs"
            elif age > 35:
                group = "35+ yrs"
            else:
                continue

            for baby in delivery.babies.filter(is_deleted=False):
                if baby.gender == 'Male':
                    age_group_summary[group]['male_count'] += 1
                elif baby.gender == 'Female':
                    age_group_summary[group]['female_count'] += 1
                age_group_summary[group]['total'] += 1

        # --------------------------------------------------
        # 7. BIRTH MODE SUMMARY
        # --------------------------------------------------
        birth_mode_summary = deliveries_qs.filter(
            no_births_to_report=False,
            birth_mode__isnull=False
        ).exclude(
            birth_mode=''
        ).values('birth_mode').annotate(
            male_count=Count('babies', filter=Q(babies__gender='Male'), distinct=True),
            female_count=Count('babies', filter=Q(babies__gender='Female'), distinct=True),
            total=Count('babies', distinct=True)
        ).order_by('birth_mode')

        # --------------------------------------------------
        # 8. TIME SLOT SUMMARY
        # --------------------------------------------------
        time_slot_summary = deliveries_qs.filter(
            no_births_to_report=False
        ).values('time_slot').annotate(
            male_count=Count('babies', filter=Q(babies__gender='Male'), distinct=True),
            female_count=Count('babies', filter=Q(babies__gender='Female'), distinct=True),
            total_in_slot=Count('babies', distinct=True)
        ).order_by('time_slot')

        # --------------------------------------------------
        # 9. FACILITY TYPE SUMMARY
        # --------------------------------------------------
        facility_type_summary = deliveries_qs.filter(
            no_births_to_report=False
        ).values('facility_type').annotate(
            male_count=Count('babies', filter=Q(babies__gender='Male'), distinct=True),
            female_count=Count('babies', filter=Q(babies__gender='Female'), distinct=True),
            total_in_type=Count('babies', distinct=True)
        ).order_by('facility_type')

        # --------------------------------------------------
        # 10. TEENAGE PREGNANCY SUMMARY
        # --------------------------------------------------
        date_10 = today.replace(year=today.year - 10)
        date_15 = today.replace(year=today.year - 15)
        date_20 = today.replace(year=today.year - 20)

        teenage_pregnancy_summary = deliveries_qs.filter(
            no_births_to_report=False,
            mother_dob__range=(date_20, date_10)
        ).values('facility').annotate(
            group_10_14=Count(
                'babies',
                filter=Q(mother_dob__range=(date_15 + timedelta(days=1), date_10)),
                distinct=True
            ),
            group_15_19=Count(
                'babies',
                filter=Q(mother_dob__range=(date_20, date_15 + timedelta(days=1))),
                distinct=True
            )
        ).order_by('facility')

        teenage_totals = teenage_pregnancy_summary.aggregate(
            total_10_14=Sum('group_10_14'),
            total_15_19=Sum('group_15_19')
        )

        # --------------------------------------------------
        # 11. MULTIPLE BIRTHS SUMMARY
        # --------------------------------------------------
        multiple_births_summary = {}
        deliveries_with_counts = deliveries_qs.annotate(
            baby_count=Count('babies', distinct=True)
        ).filter(
            baby_count__gt=1,
            no_births_to_report=False
        )

        for delivery in deliveries_with_counts:
            facility = delivery.facility
            multiple_births_summary.setdefault(
                facility,
                {'Twins': 0, 'Triplets': 0, 'Quadruplets': 0, 'Quintuplets': 0}
            )
            if delivery.baby_count == 2:
                multiple_births_summary[facility]['Twins'] += 1
            elif delivery.baby_count == 3:
                multiple_births_summary[facility]['Triplets'] += 1
            elif delivery.baby_count == 4:
                multiple_births_summary[facility]['Quadruplets'] += 1
            elif delivery.baby_count == 5:
                multiple_births_summary[facility]['Quintuplets'] += 1

        # --------------------------------------------------
        # 12. BIRTH WEIGHT SUMMARY
        # --------------------------------------------------
        weight_summary = babies_qs.aggregate(
            extremely_low=Count('id', filter=Q(weight__lt=1000)),
            very_low=Count('id', filter=Q(weight__gte=1000, weight__lt=1500)),
            low=Count('id', filter=Q(weight__gte=1500, weight__lt=2500)),
            normal=Count('id', filter=Q(weight__gte=2500, weight__lt=4000)),
            high=Count('id', filter=Q(weight__gte=4000)),
        )

        # --------------------------------------------------
        # 13. CONTEXT
        # --------------------------------------------------
        context.update({
            'form_title': "Festive Season Dashboard",

            'total_births': total_births,
            'total_males': total_males,
            'total_females': total_females,
            'total_nil_reports': total_nil_reports,

            'summary_title': title,
            'summary_table_header': header,
            'summary_footer_title': footer,
            'summary_group_by': group_by,
            'summary_data': summary_data,

            'selected_date': selected_date,
            'selected_district': selected_district,
            'selected_municipality': selected_municipality,
            'selected_facility': selected_facility,

            'district_list': [d[0] for d in DISTRICT_CHOICES if d[0]],

            'age_group_summary': age_group_summary,
            'birth_mode_summary': birth_mode_summary,
            'time_slot_summary': time_slot_summary,
            'facility_type_summary': facility_type_summary,
            'teenage_pregnancy_summary': teenage_pregnancy_summary,
            'teenage_totals': teenage_totals,
            'multiple_births_summary': multiple_births_summary,
            'has_multiple_births': bool(multiple_births_summary),
            'weight_summary': weight_summary,

            # Chart.js
            'age_group_labels': json.dumps(list(age_group_summary.keys())),
            'age_group_data': json.dumps([v['total'] for v in age_group_summary.values()]),
            'birth_mode_labels': json.dumps([i['birth_mode'] for i in birth_mode_summary]),
            'birth_mode_data': json.dumps([i['total'] for i in birth_mode_summary]),
        })

        return context


# ==========================================================
# AUTHENTICATED CRUD VIEWS
# ==========================================================
class DeliveryListView(LoginRequiredMixin, ListView):
    model = Delivery
    template_name = 'births/birth_list.html'
    context_object_name = 'deliveries'
    paginate_by = 25

    def get_queryset(self):
        user = self.request.user
        queryset = super().get_queryset().prefetch_related('babies').order_by('-timestamp')
        
        # If user is Superuser OR ProvinceUser, show all data.
        if user.is_superuser or user.groups.filter(name='ProvinceUser').exists():
            pass # No filter is applied
        # Otherwise, filter by Admin or User role
        elif user.groups.filter(name='Admin').exists():
            queryset = queryset.filter(district=user.profile.district)
        elif user.groups.filter(name='User').exists():
            queryset = queryset.filter(facility=user.profile.facility)
        else:
            return queryset.none()
        
        query = self.request.GET.get('q')
        if query:
            queryset = queryset.filter(
                Q(facility__icontains=query) | Q(mother_name__icontains=query) |
                Q(mother_surname__icontains=query) | Q(birth_mode__icontains=query)
            )
        return queryset

class DeliveryCreateView(LoginRequiredMixin, DataEditorRequiredMixin, CreateView): # <--- Added DataEditorRequiredMixin
    model = Delivery
    form_class = DeliveryForm
    template_name = 'births/birth_form.html'
    success_url = reverse_lazy('delivery_list')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form_title'] = 'Add New Delivery Record'
        context['submit_button_text'] = 'Submit Record'
        if self.request.POST:
            context['baby_formset'] = BabyFormSet(self.request.POST, prefix='babies')
        else:
            context['baby_formset'] = BabyFormSet(prefix='babies')
        return context
        
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs
        
    def form_valid(self, form):
        context = self.get_context_data()
        baby_formset = context['baby_formset']
        if form.is_valid() and baby_formset.is_valid():
            with transaction.atomic():
                form.instance.captured_by = self.request.user
                self.object = form.save()
                if not form.cleaned_data.get('no_births_to_report'):
                    # Only save babies if it's not a nil report
                    for baby_form in baby_formset:
                        if baby_form.has_changed(): 
                            baby = baby_form.save(commit=False)
                            baby.delivery = self.object
                            baby.save()
            messages.success(self.request, "Delivery record added successfully!")
            return redirect(self.get_success_url())
        messages.error(self.request, "There was an error in your submission. Please check the form.")
        return self.render_to_response(self.get_context_data(form=form))

class DeliveryUpdateView(LoginRequiredMixin, DataEditorRequiredMixin, UpdateView): # <--- Added DataEditorRequiredMixin
    model = Delivery
    form_class = DeliveryForm
    template_name = 'births/birth_form.html'
    success_url = reverse_lazy('delivery_list')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form_title'] = 'Edit Delivery Record'
        context['submit_button_text'] = 'Update Record'
        if self.request.POST:
            context['baby_formset'] = BabyFormSet(self.request.POST, instance=self.object, prefix='babies')
        else:
            context['baby_formset'] = BabyFormSet(instance=self.object, prefix='babies')
        return context
        
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs
        
    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        form = self.get_form()
        baby_formset = BabyFormSet(request.POST, instance=self.object, prefix='babies')
        if form.is_valid() and baby_formset.is_valid():
            return self.form_valid(form, baby_formset)
        else:
            print(f"--- FORM ERRORS ---\n{form.errors}\n--- FORMSET ERRORS ---\n{baby_formset.errors}\n--------------------")
            messages.error(self.request, "There was an error updating the record. Please check the form.")
            return self.form_invalid(form, baby_formset)
            
    def form_valid(self, form, baby_formset):
        with transaction.atomic(): 
            self.object = form.save()
            # If it's a nil report, ensure all babies are deleted
            if form.cleaned_data.get('no_births_to_report'):
                self.object.babies.all().delete()
            else:
                baby_formset.save() # Save babies only if not a nil report
        messages.success(self.request, "Delivery record updated successfully!")
        return redirect(self.get_success_url())
        
    def form_invalid(self, form, baby_formset):
        return self.render_to_response(self.get_context_data(form=form, baby_formset=baby_formset))

class DeliveryDeleteView(LoginRequiredMixin, DataEditorRequiredMixin, DeleteView): # <--- Added DataEditorRequiredMixin
    model = Delivery
    template_name = 'births/birth_confirm_delete.html'
    success_url = reverse_lazy('delivery_list')

# ==========================================================
# REPORTING VIEWS
# ==========================================================
@login_required
def export_full_report_excel(request):
    user = request.user
    queryset = Delivery.objects.select_related('captured_by').prefetch_related('babies').order_by('timestamp')

    if not user.is_superuser:
        if user.groups.filter(name='Admin').exists(): queryset = queryset.filter(district=user.profile.district)
        elif user.groups.filter(name='User').exists(): queryset = queryset.filter(facility=user.profile.facility)
        else: queryset = queryset.none()

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Festive Births Full Report'

    headers = [
        'Timestamp', 'Report Date', 'Time Slot', 'Time of Birth', 'District', 
        'Local Municipality', 'Facility', 'Facility Type', 
        'Mother Name', 'Mother Surname', 
        'Mother D.O.B.', 'Gravidity', 'Parity', 'Birth Mode', 'Born Before Arrival', 
        'Baby Number', 'Baby Gender', 'Baby Weight (grams)',
        'Captured By (Username)' 
    ]
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = openpyxl.styles.PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    for col_num, title in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num, value=title)
        cell.font = header_font; cell.fill = header_fill; cell.alignment = Alignment(horizontal='center')

    for delivery in queryset:
        common_data = [
            delivery.timestamp.strftime('%Y-%m-%d %H:%M'),
            delivery.report_date,
            delivery.time_slot,
            delivery.delivery_time.strftime('%H:%M') if delivery.delivery_time else '',
            delivery.district,
            delivery.local_municipality,
            delivery.facility,
            delivery.facility_type,
            delivery.mother_name,      
            delivery.mother_surname,   
            delivery.mother_dob.strftime('%Y-%m-%d') if delivery.mother_dob else '',
            delivery.gravidity,
            delivery.parity,
            delivery.birth_mode,
            'Yes' if delivery.born_before_arrival else 'No',
        ]
        
        captured_by_username = delivery.captured_by.username if delivery.captured_by else 'N/A'

        if delivery.no_births_to_report:
            sheet.append(common_data + ['NIL Report', 'N/A', 'N/A', captured_by_username])
        else:
            for i, baby in enumerate(delivery.babies.all(), 1):
                sheet.append(common_data + [i, baby.gender, baby.weight, captured_by_username])

    for col in sheet.columns:
        length = max(len(str(cell.value or '')) for cell in col)
        sheet.column_dimensions[col[0].column_letter].width = length + 2
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="festive_births_full_report.xlsx"'
    workbook.save(response)
    return response

class DashboardReportFilterView(LoginRequiredMixin, TemplateView):
    template_name = 'births/dashboard_report_filter.html'
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Pass the current user to the form's __init__ method
        context['form'] = DashboardReportFilterForm(user=self.request.user)
        context['form_title'] = 'Generate Dashboard PDF Report'
        return context

class GenerateDashboardPDF(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        # Reuse the LandingPageView to get all the calculated dashboard data.
        landing_page_view = LandingPageView()
        landing_page_view.request = request
        context = landing_page_view.get_context_data()

        context['report_user'] = request.user

        report_title = "Provincial Dashboard Report"
        if context.get('selected_facility'):
            report_title = f"{context.get('selected_facility')} Report"
        elif context.get('selected_municipality'):
            report_title = f"{context.get('selected_municipality')} Report"
        elif context.get('selected_district'):
            report_title = f"{context.get('selected_district')} Report"
        
        context['report_title'] = report_title
        
        html_string = render_to_string('births/dashboard_pdf.html', context)
        
        css_path = finders.find('css/pdf_style.css')
        if not css_path:
            return HttpResponse("Error: CSS file 'pdf_style.css' not found in static directories.", status=500)
        
        pdf_stylesheet = CSS(filename=css_path)
        
        weasyprint_html = HTML(string=html_string, base_url=request.build_absolute_uri())
        pdf_file = weasyprint_html.write_pdf(stylesheets=[pdf_stylesheet, CSS(string='@page { size: A4 portrait; }')])
        
        response = HttpResponse(pdf_file, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="dashboard_report.pdf"'
        
        return response

# --- NIL REPORTS VIEW ---
class NilReportView(LoginRequiredMixin, ListView):
    model = Delivery
    template_name = 'births/report_nil.html'
    context_object_name = 'nil_reports'
    paginate_by = 50

    def get_queryset(self):
        queryset = super().get_queryset().filter(no_births_to_report=True).select_related('captured_by')

        district = self.request.GET.get('district')
        municipality = self.request.GET.get('local_municipality')
        facility = self.request.GET.get('facility')

        if district:
            queryset = queryset.filter(district=district)
        if municipality:
            queryset = queryset.filter(local_municipality=municipality)
        if facility:
            queryset = queryset.filter(facility=facility)
            
        user = self.request.user
        if not user.is_superuser and not user.groups.filter(name='ProvinceUser').exists():
            if user.groups.filter(name='Admin').exists():
                queryset = queryset.filter(district=user.profile.district)
            elif user.groups.filter(name='User').exists():
                queryset = queryset.filter(facility=user.profile.facility)

        return queryset.order_by('-report_date', 'district', 'facility')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form'] = NilReportFilterForm(self.request.GET or None)
        context['report_title'] = "Nil Births Report"
        return context

# ==========================================================
# NEW: ABNORMAL BIRTH WEIGHT REPORT VIEW
# ==========================================================
class AbnormalWeightReportView(LoginRequiredMixin, ListView):
    model = Baby
    template_name = 'births/report_abnormal_weights.html'
    context_object_name = 'babies'
    paginate_by = 50 # Add pagination for long lists

    def get_queryset(self):
        user = self.request.user
        
        queryset = super().get_queryset().select_related('delivery', 'delivery__captured_by')

        if not user.is_superuser and not user.groups.filter(name='ProvinceUser').exists():
            if user.groups.filter(name='Admin').exists():
                queryset = queryset.filter(delivery__district=user.profile.district)
            elif user.groups.filter(name='User').exists():
                queryset = queryset.filter(delivery__facility=user.profile.facility)

        queryset = queryset.exclude(weight__gte=2500, weight__lt=4000)

        queryset = queryset.annotate(
            comment=Case(
                When(weight__lt=1000, then=Value('Extremely Low')),
                When(weight__gte=1000, weight__lt=1500, then=Value('Very Low')),
                When(weight__gte=1500, weight__lt=2500, then=Value('Low')),
                When(weight__gte=4000, then=Value('High / Macrosomic')),
                default=Value('N/A'),
                output_field=CharField(),
            )
        ).order_by('-delivery__timestamp')

        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['report_title'] = "Abnormal Birth Weight Report"
        return context

# ==========================================================
# NEW REPORTING VIEW: EXPORT USER LIST
# ==========================================================
@login_required
def export_user_list_excel(request):
    user = request.user

    if not user.is_superuser:
        messages.error(request, "You do not have permission to export the user list.")
        return redirect('landing_page') 

    queryset = User.objects.select_related('profile').order_by('first_name', 'last_name')

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'User List Report'

    headers = [
        'Name', 'Surname', 'Email', 'Title', 'Designation',
        'Persal Number', 'Mobile Number', 'Role(s)',
        'Allocated District', 'Allocated Local Municipality', 'Allocated Facility',
        'Active Account', 'Superuser Status'
    ]
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = openpyxl.styles.PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    for col_num, title in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for app_user in queryset:
        user_profile = None
        try:
            user_profile = app_user.profile 
        except Profile.DoesNotExist:
            pass

        roles_str = ", ".join([group.name for group in app_user.groups.all()])

        row_data = [
            app_user.first_name,
            app_user.last_name,
            app_user.email,
            user_profile.title if user_profile else '',
            user_profile.designation if user_profile else '',
            user_profile.persal_number if user_profile else '',
            user_profile.mobile_number if user_profile else '',
            roles_str,
            user_profile.district if user_profile else '',
            user_profile.local_municipality if user_profile else '',
            user_profile.facility if user_profile else '',
            'Yes' if app_user.is_active else 'No', 
            'Yes' if app_user.is_superuser else 'No',
        ]
        sheet.append(row_data)

    for col in sheet.columns:
        length = max(len(str(cell.value or '')) for cell in col)
        sheet.column_dimensions[col[0].column_letter].width = length + 2
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="user_list_report.xlsx"'
    workbook.save(response)
    return response

# ==========================================================
# AJAX HELPER VIEWS
# ==========================================================
def load_options(request):
    parent_type = request.GET.get('type')
    parent_id = request.GET.get('id')
    
    options_list = []
    if parent_type == 'district' and parent_id:
        options_list = LOCATION_DATA['municipalities'].get(parent_id, [])
    elif parent_type == 'municipality' and parent_id:
        options_list = LOCATION_DATA['facilities'].get(parent_id, [])
        
    return JsonResponse({'options': options_list})

def get_facility_type(request):
    facility_name = request.GET.get('facility_name')
    if not facility_name:
        return JsonResponse({'error': 'No facility name provided'}, status=400)

    facility_type = FACILITY_TYPES.get(facility_name)
    
    if facility_type is None:
        return JsonResponse({'error': 'Facility type not found'}, status=404)
        
    return JsonResponse({'facility_type': facility_type})

